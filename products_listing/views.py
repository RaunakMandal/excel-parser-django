from django.shortcuts import render
from openpyxl import load_workbook
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponseRedirect
from .models import Product, ProductVariation
from django.core.paginator import Paginator
import random
import datetime

def redirect(request):
    return HttpResponseRedirect("/products")

# This is the Main method which will be called when the user visits the /products page.
def getProducts(request):
    # Get the page number and search query from url parameters or set them to default values
    page = request.GET.get('page') or 1
    q = request.GET.get('q') or ''

    # Get the sort column and sort direction from url parameters or set them to default values
    sortcol = request.GET.get('sortcol') or 'name'
    sortdir = request.GET.get('sortdir') or 'asc'

    # Get the products from the database, including filtering and sorting
    products = Product.objects.filter(name__icontains=q).order_by(sortcol).reverse() if sortdir == 'desc' else Product.objects.filter(name__icontains=q).order_by(sortcol)
    for product in products:
        product_variations = ProductVariation.objects.filter(product_id=product.uid)
        product.variations = product_variations

    # Paginate the products to show 5 products per page and finally return the paginated products
    paginator = Paginator(products, 5)
    context = {
        'products': paginator.get_page(page),
    }

    return render(request, 'products_listing/main.html', context)


# This method will be called when the user uploads an excel file

# Currently we are not using the csrf token, so it is exempted for now.
@csrf_exempt
def add_product(request):
    # Check if the request is a POST request and if the request has a file
    if request.method == 'POST' and request.FILES['excel']:
        file = request.FILES['excel']

        # Check if the file is of type XLSX/XLS, else return an error
        if file.content_type not in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
            return JsonResponse({'success': False, 'message': 'Invalid file type'})

        # Load the workbook and get the first sheet for our use case
        workbook = load_workbook(filename=file)
        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]

        for i, row in enumerate(worksheet.iter_rows()):
            if i == 0:
                continue

            products = Product.objects.filter(name=row[0].value)
            # Check if the product already exists in the database
            if products:
                # If the product exists, check if the variation already exists
                for product in products:
                    product_variations_list = ProductVariation.objects.filter(product_id=product.uid, variation_text=row[1].value)

                    # If the variation exists, update the stock
                    if product_variations_list:
                        for product_variation in product_variations_list:
                            product_variation.stock += row[2].value
                            product_variation.save()
                            
                            product.updated_at = datetime.datetime.now()
                            product.save()

                    # If the variation does not exist, create a new variation
                    else:
                        ProductVariation.objects.create(
                            product_id=product,
                            variation_text=row[1].value,
                            stock=row[2].value,
                        )
                        product.updated_at = datetime.datetime.now()
                        product.save()

            # If the product does not exist, create a new product, variation and set a random lowest price
            else:
                # As we did not have any lowest price, putting an arbitrary value here.
                price = [40000, 34999, 32000, 20080, 60000]
                product = Product.objects.create(
                    name=row[0].value,
                    lowest_price=price[random.randint(0, 4)],
                )
                product.save()
                ProductVariation.objects.create(
                    product_id=product,
                    variation_text=row[1].value,
                    stock=row[2].value,
                ).save()

        return JsonResponse({'success': True, 'message': 'Product added successfully'})
    return JsonResponse({'success': False, 'message': 'Invalid request'})